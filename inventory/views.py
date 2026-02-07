import json
import re
import unicodedata
import xml.etree.ElementTree as ET
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import F, Sum
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views import View

from core.utils import get_user_company
from p_v_App.models import Category, Estoque, Products

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


def _normalize_header(value):
    normalized = (
        unicodedata.normalize('NFKD', str(value or '')).encode(
            'ASCII', 'ignore').decode('ASCII')
    )
    return normalized.strip().lower()


def _parse_status_cell(raw_value):
    if isinstance(raw_value, (int, float)):
        int_value = int(raw_value)
        if int_value in (0, 1):
            return int_value
    text = str(raw_value or '').strip().lower()
    mapping = {
        '1': 1,
        'ativo': 1,
        'active': 1,
        'sim': 1,
        'yes': 1,
        '0': 0,
        'inativo': 0,
        'inactive': 0,
        'nao': 0,
        'não': 0,
        'no': 0,
    }
    return mapping.get(text)


def _parse_decimal_cell(raw_value):
    if raw_value is None:
        return None
    if isinstance(raw_value, Decimal):
        return raw_value
    if isinstance(raw_value, (int, float)):
        return Decimal(str(raw_value))

    text = str(raw_value).strip()
    if not text:
        return None

    cleaned = re.sub(r'[^0-9,.-]', '', text)
    if not cleaned:
        return None

    if ',' in cleaned and '.' in cleaned:
        if cleaned.rfind(',') > cleaned.rfind('.'):
            cleaned = cleaned.replace('.', '')
            cleaned = cleaned.replace(',', '.')
        else:
            cleaned = cleaned.replace(',', '')
    elif ',' in cleaned:
        cleaned = cleaned.replace(',', '.')

    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None


def _parse_int_cell(raw_value):
    decimal_value = _parse_decimal_cell(raw_value)
    if decimal_value is None:
        return None
    try:
        return int(decimal_value.to_integral_value(rounding=ROUND_HALF_UP))
    except (InvalidOperation, ValueError):
        return None


def _parse_validade_cell(raw_value):
    allowed = {choice[0] for choice in Estoque.VALIDADE_CHOICES}
    if raw_value is None:
        return 0 if 0 in allowed else None
    if isinstance(raw_value, (int, float, Decimal)):
        value = int(float(raw_value))
    else:
        text = str(raw_value).strip().lower()
        if not text:
            return 0 if 0 in allowed else None
        mapping = {
            'sem validade': 0,
            '30 dias': 30,
            '60 dias': 60,
            '90 dias': 90,
            '120 dias': 120,
            '180 dias': 180,
            '365 dias': 365,
        }
        if text in mapping:
            value = mapping[text]
        else:
            digits = re.sub(r'[^0-9-]', '', text)
            if not digits:
                return None
            value = int(digits)

    return value if value in allowed else None


@login_required
def estoque(request):
    user_company = get_user_company(request)
    if not user_company:
        messages.error(
            request, 'Usuário não está associado a nenhuma empresa.')
        return render(request, 'inventory/estoque.html', {'estoque': []})

    query = request.GET.get('q', '').strip()
    page = request.GET.get('page', 1)

    base_qs = Estoque.objects.filter(company=user_company)
    estoque_qs = base_qs.filter(
        produto__name__icontains=query) if query else base_qs
    estoque_qs = estoque_qs.order_by('-id')

    paginator = Paginator(estoque_qs, 30)
    try:
        estoque_paginated = paginator.page(page)
    except PageNotAnInteger:
        estoque_paginated = paginator.page(1)
    except EmptyPage:
        estoque_paginated = paginator.page(paginator.num_pages)

    total_items = base_qs.aggregate(total=Sum('quantidade'))['total'] or 0
    total_value = base_qs.aggregate(
        total=Sum(F('quantidade') * F('produto__price')))['total'] or Decimal('0')
    total_cost = base_qs.aggregate(
        total=Sum(F('quantidade') * F('produto__custo')))['total'] or Decimal('0')

    context = {
        'page_title': 'Lista de Produtos',
        'estoque': estoque_paginated,
        'q': query,
        'total_items': int(total_items) if total_items else 0,
        'total_value': float(total_value) if total_value else 0.0,
        'total_cost': float(total_cost) if total_cost else 0.0,
    }
    return render(request, 'inventory/estoque.html', context)


@login_required
def delete_product_estoque(request):
    data = request.POST
    resp = {'status': 'failed'}
    user_company = get_user_company(request)
    if not user_company:
        resp['msg'] = 'Usuário não está associado a nenhuma empresa.'
        return JsonResponse(resp)

    estoque_id = data.get('id')
    try:
        estoque = Estoque.objects.get(pk=estoque_id, company=user_company)
    except Estoque.DoesNotExist:
        resp['msg'] = 'Produto não encontrado.'
        return JsonResponse(resp)

    estoque.delete()
    messages.success(request, 'Produto deletado com sucesso.')
    resp['status'] = 'success'
    return JsonResponse(resp)


@login_required
def manage_products_estoque(request):
    user_company = get_user_company(request)
    if not user_company:
        messages.error(
            request, 'Usuário não está associado a nenhuma empresa.')
        return render(
            request,
            'inventory/manage_estoque.html',
            {'product': None, 'categories': [], 'products': []},
        )

    product = None
    prod_id = request.GET.get('id', '').strip()
    if prod_id.isnumeric():
        product = get_object_or_404(
            Estoque, pk=int(prod_id), company=user_company)

    categories = Category.objects.filter(
        status=1, company=user_company).order_by('name')
    products = Products.objects.filter(company=user_company).order_by('name')

    return render(
        request,
        'inventory/manage_estoque.html',
        {'product': product, 'categories': categories, 'products': products},
    )


@login_required
def save_product_estoque(request):
    data = request.POST
    resp = {'status': 'failed', 'msg': ''}

    user_company = get_user_company(request)
    if not user_company:
        resp['msg'] = 'Usuário não está associado a nenhuma empresa.'
        return JsonResponse(resp)

    record_id = data.get('id', '').strip()
    produto_id = data.get('produto_id', '').strip()
    categoria_id = data.get('categoria_id', '').strip()
    quantidade = data.get('quantidade', '0').strip()
    validade = data.get('validade', '0').strip()
    preco = data.get('price', '0').replace(',', '.').strip()
    custo = data.get('custo', '0').replace(',', '.').strip()
    status = data.get('status', '1').strip()

    try:
        produto = get_object_or_404(
            Products, pk=int(produto_id), company=user_company)
        categoria = get_object_or_404(
            Category, pk=int(categoria_id), company=user_company)
    except (ValueError, Products.DoesNotExist, Category.DoesNotExist):
        resp['msg'] = 'Produto ou categoria inválido.'
        return JsonResponse(resp)

    if record_id.isnumeric() and int(record_id) > 0:
        estoque = get_object_or_404(
            Estoque, pk=int(record_id), company=user_company)
    else:
        estoque = Estoque(company=user_company)

    estoque.produto = produto
    estoque.categoria = categoria
    estoque.quantidade = int(quantidade) if quantidade.isnumeric() else 0
    estoque.validade = int(validade) if validade.isnumeric() else 0
    estoque.preco = preco
    estoque.custo = custo
    estoque.status = int(status) if status in ('0', '1') else 1

    try:
        estoque.save()
        messages.success(request, 'Produto salvo com sucesso.')
        return JsonResponse({'status': 'success'})
    except Exception as exc:
        resp['msg'] = f'Erro ao salvar: {exc}'
        return JsonResponse(resp)


@login_required
def upload_estoque(request):
    user_company = get_user_company(request)
    if not user_company:
        messages.error(
            request, 'Usuário não está associado a nenhuma empresa.')
        return redirect('estoque')

    expected_headers = [
        'Código do Produto',
        'Nome do Produto',
        'Categoria',
        'Quantidade',
        'Validade (dias)',
        'Preço',
        'Custo',
        'Status',
    ]

    if request.method == 'GET':
        return render(
            request,
            'inventory/upload_estoque.html',
            {'expected_headers': expected_headers},
        )

    resp = {'status': 'failed'}
    upload_file = request.FILES.get('file')
    if not upload_file:
        resp['msg'] = 'Selecione um arquivo Excel (.xlsx) para importar.'
        return JsonResponse(resp)

    try:
        workbook = load_workbook(upload_file)
    except Exception:
        resp['msg'] = 'Não foi possível ler o arquivo enviado. Utilize um arquivo .xlsx válido.'
        return JsonResponse(resp)

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        resp['msg'] = 'O arquivo enviado está vazio.'
        return JsonResponse(resp)

    header_row = rows[0]
    normalized_header = [_normalize_header(value) for value in header_row]
    header_map = {value: index for index,
                  value in enumerate(normalized_header) if value}

    def resolve_index(possible_keys):
        for key in possible_keys:
            if key in header_map:
                return header_map[key]
        return None

    code_idx = resolve_index(
        ['codigo', 'código', 'codigo do produto', 'código do produto', 'sku'])
    category_idx = resolve_index(['categoria', 'categoria do produto'])
    quantity_idx = resolve_index(['quantidade', 'qtd', 'estoque'])
    validity_idx = resolve_index(
        ['validade', 'validade (dias)', 'validade em dias'])
    price_idx = resolve_index(['preco', 'preço', 'valor', 'valor de venda'])
    cost_idx = resolve_index(['custo', 'custo unitario', 'custo unitário'])
    status_idx = resolve_index(['status'])

    if None in (code_idx, quantity_idx, status_idx):
        resp['msg'] = 'Cabeçalho inválido. Utilize o modelo de importação disponibilizado.'
        return JsonResponse(resp)

    created_count = 0
    updated_count = 0
    error_rows = []

    def is_empty_row(row):
        return all(
            (
                cell is None
                or (isinstance(cell, str) and not cell.strip())
                or (not isinstance(cell, str) and str(cell).strip() == '')
            )
            for cell in row
        )

    for row_number, row in enumerate(rows[1:], start=2):
        if not row or is_empty_row(row):
            continue

        code_cell = row[code_idx] if len(row) > code_idx else None
        category_cell = row[category_idx] if category_idx is not None and len(
            row) > category_idx else None
        quantity_cell = row[quantity_idx] if len(row) > quantity_idx else None
        validity_cell = row[validity_idx] if validity_idx is not None and len(
            row) > validity_idx else None
        price_cell = row[price_idx] if price_idx is not None and len(
            row) > price_idx else None
        cost_cell = row[cost_idx] if cost_idx is not None and len(
            row) > cost_idx else None
        status_cell = row[status_idx] if len(row) > status_idx else None

        code = str(code_cell or '').strip()
        category_name = str(category_cell or '').strip()
        quantity_value = _parse_int_cell(quantity_cell)
        validity_value = _parse_validade_cell(validity_cell)
        status_value = _parse_status_cell(status_cell)
        price_value = _parse_decimal_cell(price_cell)
        cost_value = _parse_decimal_cell(cost_cell)

        if not code:
            error_rows.append(
                f'Linha {row_number}: o código do produto é obrigatório.')
            continue

        if quantity_value is None:
            error_rows.append(
                f'Linha {row_number}: informe uma quantidade válida.')
            continue

        if status_value is None:
            error_rows.append(
                f"Linha {row_number}: status inválido. Utilize 'Ativo' ou 'Inativo' (ou 1/0)."
            )
            continue

        if validity_value is None:
            error_rows.append(
                f'Linha {row_number}: validade inválida. Utilize um dos valores permitidos (0, 30, 60, 90, 120, 180, 365).'
            )
            continue

        product = Products.objects.filter(
            company=user_company, code__iexact=code).first()
        if not product:
            error_rows.append(
                f"Linha {row_number}: produto com código '{code}' não encontrado.")
            continue

        category = None
        if category_name:
            category = Category.objects.filter(
                company=user_company, name__iexact=category_name).first()
            if not category:
                category = Category.objects.create(
                    company=user_company,
                    name=category_name,
                    description='',
                    status=1,
                )
        else:
            category = product.category_id

        if not category:
            error_rows.append(
                f"Linha {row_number}: não foi possível determinar a categoria para o produto '{code}'."
            )
            continue

        try:
            estoque_obj = Estoque.objects.filter(
                company=user_company, produto=product).first()
            if estoque_obj:
                estoque_obj.categoria = category
                estoque_obj.quantidade = quantity_value
                estoque_obj.validade = validity_value
                if price_value is not None:
                    estoque_obj.preco = float(price_value)
                elif estoque_obj.preco in (None, 0) and product.price is not None:
                    estoque_obj.preco = float(product.price)
                if cost_value is not None:
                    estoque_obj.custo = float(cost_value)
                elif estoque_obj.custo in (None, 0) and product.custo is not None:
                    estoque_obj.custo = float(product.custo)
                estoque_obj.descricao = product
                estoque_obj.status = status_value
                estoque_obj.save()
                updated_count += 1
            else:
                Estoque.objects.create(
                    company=user_company,
                    produto=product,
                    categoria=category,
                    quantidade=quantity_value,
                    validade=validity_value,
                    preco=float(price_value) if price_value is not None else float(
                        product.price or 0),
                    custo=float(cost_value) if cost_value is not None else float(
                        product.custo or 0),
                    status=status_value,
                    descricao=product,
                )
                created_count += 1
        except Exception as exc:
            error_rows.append(
                f'Linha {row_number}: erro ao salvar estoque ({exc}).')
            continue

    if created_count or updated_count:
        if error_rows:
            messages.warning(
                request,
                (
                    f'Importação concluída: {created_count} item(ns) adicionados e '
                    f'{updated_count} atualizados. Algumas linhas foram ignoradas.'
                ),
            )
        else:
            messages.success(
                request,
                (
                    f'Importação concluída: {created_count} item(ns) adicionados e '
                    f'{updated_count} atualizados.'
                ),
            )
    else:
        if error_rows:
            messages.error(
                request,
                'Nenhum item de estoque foi importado. Verifique o arquivo enviado.',
            )

    if error_rows:
        resp['status'] = 'partial'
        resp['errors'] = error_rows
        if created_count or updated_count:
            resp['msg'] = (
                'Importação concluída com pendências. '
                f'{created_count} item(ns) adicionados e '
                f'{updated_count} atualizados.'
            )
        else:
            resp['msg'] = (
                'Nenhum item foi importado. Revise as pendências indicadas no arquivo.'
            )
    else:
        resp['status'] = 'success'
        resp['msg'] = (
            'Estoque importado com sucesso. '
            f'{created_count} adicionados e {updated_count} atualizados.'
        )

    resp['created'] = created_count
    resp['updated'] = updated_count

    return JsonResponse(resp)


@login_required
def download_estoque_template(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Estoque'

    headers = [
        'Código do Produto',
        'Nome do Produto',
        'Categoria',
        'Quantidade',
        'Validade (dias)',
        'Preço',
        'Custo',
        'Status',
    ]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    worksheet.append([
        'PRD-0001',
        'Refrigerante 350ml',
        'Bebidas',
        24,
        90,
        6.5,
        3.2,
        'Ativo',
    ])
    worksheet.append([
        'PRD-0002',
        'Pizza Família',
        'Alimentos',
        5,
        0,
        45.0,
        22.5,
        'Inativo',
    ])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return FileResponse(
        buffer,
        as_attachment=True,
        filename='modelo_importacao_estoque.xlsx',
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


class EstoqueXMLUploadView(View):
    template_name = 'inventory/estoque_xml_upload.html'

    def get(self, request):
        user_company = get_user_company(request)
        if not user_company:
            messages.error(
                request, 'Usuário não está associado a nenhuma empresa.')
            return redirect('estoque')

        return render(
            request,
            self.template_name,
            {
                'expected_fields': ['cProd', 'xProd', 'qCom', 'vUnCom', 'vUnTrib'],
                'csrf_token_value': request.META.get('CSRF_COOKIE', ''),
            },
        )

    def post(self, request):
        user_company = get_user_company(request)
        if not user_company:
            return JsonResponse(
                {'status': 'failed',
                 'msg': 'Usuário não está associado a nenhuma empresa.'}
            )

        content_type = request.META.get('CONTENT_TYPE', '')
        if content_type.startswith('application/json'):
            try:
                payload = json.loads(request.body.decode('utf-8'))
                items = payload.get('items') or []
            except (TypeError, ValueError):
                return JsonResponse(
                    {'status': 'failed', 'msg': 'Payload JSON inválido.'}
                )
            return self._apply_items(user_company, items)

        upload_file = request.FILES.get('file')
        if not upload_file:
            return JsonResponse(
                {'status': 'failed', 'msg': 'Envie um arquivo XML para importar.'}
            )

        parsed_items, errors = self._parse_xml_items(upload_file, user_company)
        status = 'success' if parsed_items else 'failed'
        if errors and parsed_items:
            status = 'partial'
        if not parsed_items and not errors:
            errors.append('Nenhum item de estoque encontrado no XML informado.')

        return JsonResponse(
            {
                'status': status,
                'items': parsed_items,
                'errors': errors,
                'msg': 'Pré-visualização pronta.' if parsed_items else 'Nenhum item encontrado.',
            }
        )

    def _parse_xml_items(self, upload_file, company):
        try:
            tree = ET.parse(upload_file)
            root = tree.getroot()
        except Exception:
            return [], ['Não foi possível ler o XML enviado.']

        def strip_tag(tag):
            return tag.split('}', 1)[-1] if '}' in tag else tag

        def find_child_text(element, name):
            for child in element:
                if strip_tag(child.tag) == name:
                    return (child.text or '').strip()
            return ''

        items = []
        errors = []
        for node in root.iter():
            if strip_tag(node.tag) != 'det':
                continue

            prod = next(
                (child for child in node if strip_tag(child.tag) == 'prod'),
                None,
            )
            if prod is None:
                continue

            code = find_child_text(prod, 'cProd')
            name = find_child_text(prod, 'xProd')
            if not code and not name:
                continue

            quantity_value = _parse_decimal_cell(find_child_text(prod, 'qCom')) or Decimal('0')

            xml_unit_price = _parse_decimal_cell(
                find_child_text(prod, 'vUnCom') or find_child_text(prod, 'vProd')
            )
            xml_unit_cost = _parse_decimal_cell(find_child_text(prod, 'vUnTrib'))

            product = Products.objects.filter(
                company=company, code__iexact=code).first() if code else None
            category_name = ''
            if product and product.category_id:
                category_name = product.category_id.name

            product_price = _parse_decimal_cell(getattr(product, 'price', None)) if product else None
            product_cost = _parse_decimal_cell(getattr(product, 'custo', None)) if product else None

            cost_value = xml_unit_price if xml_unit_price is not None else xml_unit_cost
            if cost_value is None:
                cost_value = product_cost

            price_value = product_price if product_price is not None else (xml_unit_price or xml_unit_cost)

            items.append(
                {
                    'code': code,
                    'name': name,
                    'quantity': float(quantity_value) if quantity_value is not None else 0,
                    'price': float(price_value)
                    if price_value is not None
                    else float(product.price if product else 0),
                    'cost': float(cost_value)
                    if cost_value is not None
                    else float(product.custo if product else 0),
                    'category': category_name,
                    'status': 1,
                }
            )

        if not items:
            errors.append('Nenhum item de estoque encontrado no XML informado.')

        return items, errors

    def _apply_items(self, company, items):
        if not isinstance(items, list):
            return JsonResponse({'status': 'failed', 'msg': 'Lista de itens inválida.'})

        created = 0
        updated = 0
        errors: list[str] = []

        for idx, item in enumerate(items, start=1):
            code = str(item.get('code') or '').strip()
            if not code:
                errors.append(f'Linha {idx}: informe o código do produto.')
                continue

            qty_value = _parse_int_cell(item.get('quantity'))
            if qty_value is None or qty_value < 0:
                errors.append(f'Linha {idx}: quantidade inválida.')
                continue

            status_value = _parse_status_cell(item.get('status'))
            if status_value is None:
                status_value = 1

            price_value = _parse_decimal_cell(item.get('price'))
            cost_value = _parse_decimal_cell(item.get('cost'))

            product = Products.objects.filter(
                company=company, code__iexact=code).first()
            if not product:
                errors.append(f'Linha {idx}: produto {code} não encontrado.')
                continue

            category = product.category_id
            category_name = str(item.get('category') or '').strip()
            if category_name:
                category = Category.objects.filter(
                    company=company, name__iexact=category_name
                ).first()
                if not category:
                    category = Category.objects.create(
                        company=company,
                        name=category_name,
                        description='',
                        status=1,
                    )

            estoque_obj = Estoque.objects.filter(
                company=company, produto=product).first()
            if estoque_obj:
                estoque_obj.quantidade = (estoque_obj.quantidade or 0) + qty_value
                estoque_obj.categoria = category
                if price_value is not None:
                    estoque_obj.preco = float(price_value)
                if cost_value is not None:
                    estoque_obj.custo = float(cost_value)
                estoque_obj.descricao = product
                estoque_obj.status = status_value if status_value in (0, 1) else estoque_obj.status
                estoque_obj.save()
                updated += 1
            else:
                Estoque.objects.create(
                    company=company,
                    produto=product,
                    categoria=category,
                    quantidade=qty_value,
                    validade=0,
                    preco=float(price_value if price_value is not None else product.price or 0),
                    custo=float(cost_value if cost_value is not None else product.custo or 0),
                    status=status_value if status_value in (0, 1) else 1,
                    descricao=product,
                )
                created += 1

        status = 'success' if not errors else 'partial'
        msg = f'Estoque atualizado: {created} criado(s) e {updated} atualizado(s).'
        if not created and not updated:
            status = 'failed'
            msg = 'Nenhum item foi aplicado.'

        return JsonResponse(
            {
                'status': status,
                'created': created,
                'updated': updated,
                'errors': errors,
                'msg': msg,
            }
        )
