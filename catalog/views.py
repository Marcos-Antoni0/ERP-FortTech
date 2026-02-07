import json
import re
import unicodedata
import xml.etree.ElementTree as ET
from io import BytesIO
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Avg, ProtectedError, Q, Sum
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.views import View

from core.utils import get_user_company
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from p_v_App.models import Category, ProductComboItem, Products


@login_required
def category(request):
    user_company = get_user_company(request)
    if not user_company:
        messages.error(
            request, 'Usuário não está associado a nenhuma empresa.')
        return redirect('home-page')

    query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '').strip()
    page = request.GET.get('page', 1)

    base_qs = Category.objects.filter(company=user_company)
    if query:
        base_qs = base_qs.filter(name__icontains=query)
    if status_filter in ['0', '1']:
        base_qs = base_qs.filter(status=int(status_filter))

    category_qs = base_qs.order_by('-id')
    paginator = Paginator(category_qs, 20)
    try:
        category_paginated = paginator.page(page)
    except PageNotAnInteger:
        category_paginated = paginator.page(1)
    except EmptyPage:
        category_paginated = paginator.page(paginator.num_pages)

    context = {
        'page_title': 'Lista de Categorias',
        'category': category_paginated,
        'q': query,
        'status_filter': status_filter,
        'total_categories': base_qs.count(),
        'active_categories': base_qs.filter(status=1).count(),
        'inactive_categories': base_qs.filter(status=0).count(),
    }
    return render(request, 'catalog/category.html', context)


@login_required
def manage_category(request):
    user_company = get_user_company(request)
    if not user_company:
        messages.error(
            request, 'Usuário não está associado a nenhuma empresa.')
        return redirect('category-page')

    category_obj = {}
    if request.method == 'GET':
        category_id = request.GET.get('id', '')
        if category_id.isnumeric() and int(category_id) > 0:
            category_obj = Category.objects.filter(
                id=category_id, company=user_company).first()

    return render(request, 'catalog/manage_category.html', {'category': category_obj})


@login_required
def save_category(request):
    data = request.POST
    resp = {'status': 'failed'}
    try:
        user_company = get_user_company(request)
        if not user_company:
            resp['msg'] = 'Usuário não está associado a nenhuma empresa.'
            return HttpResponse(json.dumps(resp), content_type='application/json')

        category_id = data.get('id', '')
        if category_id.isnumeric() and int(category_id) > 0:
            Category.objects.filter(id=category_id, company=user_company).update(
                name=data.get('name', ''),
                description=data.get('description', ''),
                status=data.get('status', 1),
            )
        else:
            Category.objects.create(
                name=data.get('name', ''),
                description=data.get('description', ''),
                status=data.get('status', 1),
                company=user_company,
            )
        resp['status'] = 'success'
        messages.success(request, 'Categoria salva com sucesso.')
    except Exception as exc:
        resp['msg'] = str(exc)
    return HttpResponse(json.dumps(resp), content_type='application/json')


@login_required
def delete_category(request):
    data = request.POST
    resp = {'status': ''}
    user_company = get_user_company(request)
    if not user_company:
        resp['status'] = 'failed'
        resp['msg'] = 'Usuário não está associado a nenhuma empresa.'
        return HttpResponse(json.dumps(resp), content_type='application/json')

    try:
        Category.objects.filter(id=data.get(
            'id'), company=user_company).delete()
        resp['status'] = 'success'
        messages.success(request, 'Categoria deletada com sucesso.')
    except Exception as exc:
        resp['status'] = 'failed'
        resp['msg'] = f'Erro ao deletar categoria: {exc}'
    return HttpResponse(json.dumps(resp), content_type='application/json')


def _normalize_header(value):
    normalized = unicodedata.normalize('NFKD', str(value or '')).encode(
        'ASCII', 'ignore').decode('ASCII')
    return normalized.strip().lower()


def _parse_status_cell(raw_value):
    if isinstance(raw_value, (int, float)):
        int_value = int(raw_value)
        if int_value in (0, 1):
            return int_value
    text = str(raw_value or '').strip().lower()
    mapping = {
        '1': 1,
        'ativo': 1,
        'active': 1,
        'sim': 1,
        'yes': 1,
        '0': 0,
        'inativo': 0,
        'inactive': 0,
        'nao': 0,
        'não': 0,
        'no': 0,
    }
    return mapping.get(text)


def _parse_decimal_cell(raw_value):
    if raw_value is None:
        return None
    if isinstance(raw_value, Decimal):
        return raw_value
    if isinstance(raw_value, (int, float)):
        return Decimal(str(raw_value))

    text = str(raw_value).strip()
    if not text:
        return None

    cleaned = re.sub(r'[^0-9,.-]', '', text)
    if not cleaned:
        return None

    if ',' in cleaned and '.' in cleaned:
        if cleaned.rfind(',') > cleaned.rfind('.'):
            cleaned = cleaned.replace('.', '')
            cleaned = cleaned.replace(',', '.')
        else:
            cleaned = cleaned.replace(',', '')
    elif ',' in cleaned:
        cleaned = cleaned.replace(',', '.')

    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None


@login_required
def upload_categories(request):
    user_company = get_user_company(request)
    if not user_company:
        messages.error(
            request, 'Usuário não está associado a nenhuma empresa.')
        return redirect('category-page')

    expected_headers = ['Nome', 'Descrição', 'Status']

    if request.method == 'GET':
        return render(
            request,
            'catalog/upload_categories.html',
            {'expected_headers': expected_headers},
        )

    resp = {'status': 'failed'}
    upload_file = request.FILES.get('file')
    if not upload_file:
        resp['msg'] = 'Selecione um arquivo Excel (.xlsx) para importar.'
        return JsonResponse(resp)

    try:
        workbook = load_workbook(upload_file)
    except Exception:
        resp['msg'] = 'Não foi possível ler o arquivo enviado. Utilize um arquivo .xlsx válido.'
        return JsonResponse(resp)

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        resp['msg'] = 'O arquivo enviado está vazio.'
        return JsonResponse(resp)

    header_row = rows[0]
    normalized_header = [_normalize_header(value) for value in header_row]
    header_map = {value: index for index,
                  value in enumerate(normalized_header) if value}

    def resolve_index(possible_keys):
        for key in possible_keys:
            if key in header_map:
                return header_map[key]
        return None

    name_idx = resolve_index(['nome'])
    description_idx = resolve_index(['descricao', 'descrição'])
    status_idx = resolve_index(['status'])

    if None in (name_idx, description_idx, status_idx):
        resp['msg'] = 'Cabeçalho inválido. Utilize o modelo de importação disponibilizado.'
        return JsonResponse(resp)

    created_count = 0
    updated_count = 0
    error_rows = []

    def is_empty_row(row):
        return all(
            (
                cell is None
                or (isinstance(cell, str) and not cell.strip())
                or (not isinstance(cell, str) and str(cell).strip() == '')
            )
            for cell in row
        )

    for row_number, row in enumerate(rows[1:], start=2):
        if not row or is_empty_row(row):
            continue

        name_cell = row[name_idx] if len(row) > name_idx else None
        description_cell = row[description_idx] if len(
            row) > description_idx else None
        status_cell = row[status_idx] if len(row) > status_idx else None

        name = str(name_cell or '').strip()
        description = str(description_cell or '').strip()

        if not name:
            error_rows.append(
                f'Linha {row_number}: o nome da categoria é obrigatório.')
            continue

        status_value = _parse_status_cell(status_cell)
        if status_value is None:
            error_rows.append(
                f"Linha {row_number}: status inválido. Utilize 'Ativo' ou 'Inativo' (ou 1/0)."
            )
            continue

        try:
            _, created = Category.objects.update_or_create(
                company=user_company,
                name=name,
                defaults={'description': description, 'status': status_value},
            )
        except Exception as exc:
            error_rows.append(
                f'Linha {row_number}: erro ao salvar categoria ({exc}).')
            continue

        if created:
            created_count += 1
        else:
            updated_count += 1

    if created_count or updated_count:
        if error_rows:
            messages.warning(
                request,
                (
                    f'Importação concluída: {created_count} categoria(s) adicionada(s) e '
                    f'{updated_count} atualizada(s). Algumas linhas foram ignoradas.'
                ),
            )
        else:
            messages.success(
                request,
                (
                    f'Importação concluída: {created_count} categoria(s) adicionada(s) e '
                    f'{updated_count} atualizada(s).'
                ),
            )
    else:
        if error_rows:
            messages.error(
                request,
                'Nenhuma categoria foi importada. Verifique o arquivo enviado.',
            )

    if error_rows:
        resp['status'] = 'partial'
        resp['errors'] = error_rows
        if created_count or updated_count:
            resp['msg'] = (
                'Importação concluída com pendências. '
                f'{created_count} categoria(s) adicionada(s) e '
                f'{updated_count} atualizada(s).'
            )
        else:
            resp['msg'] = (
                'Nenhuma categoria foi importada. Revise as pendências indicadas no arquivo.'
            )
    else:
        resp['status'] = 'success'
        resp['msg'] = (
            'Categorias importadas com sucesso. '
            f'{created_count} adicionada(s) e {updated_count} atualizada(s).'
        )

    resp['created'] = created_count
    resp['updated'] = updated_count

    return JsonResponse(resp)


@login_required
def download_category_template(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Categorias'

    headers = ['Nome', 'Descrição', 'Status']
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    worksheet.append(
        ['Bebidas', 'Produtos e bebidas prontas para venda', 'Ativo'])
    worksheet.append(
        ['Serviços', 'Serviços temporariamente indisponíveis', 'Inativo'])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return FileResponse(
        buffer,
        as_attachment=True,
        filename='modelo_importacao_categorias.xlsx',
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@login_required
def upload_products(request):
    user_company = get_user_company(request)
    if not user_company:
        messages.error(
            request, 'Usuário não está associado a nenhuma empresa.')
        return redirect('product-page')

    expected_headers = [
        'Código',
        'Nome',
        'Descrição',
        'Categoria',
        'Preço',
        'Custo',
        'Status',
    ]

    if request.method == 'GET':
        return render(
            request,
            'catalog/upload_products.html',
            {'expected_headers': expected_headers},
        )

    resp = {'status': 'failed'}
    upload_file = request.FILES.get('file')
    if not upload_file:
        resp['msg'] = 'Selecione um arquivo Excel (.xlsx) para importar.'
        return JsonResponse(resp)

    try:
        workbook = load_workbook(upload_file)
    except Exception:
        resp['msg'] = 'Não foi possível ler o arquivo enviado. Utilize um arquivo .xlsx válido.'
        return JsonResponse(resp)

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        resp['msg'] = 'O arquivo enviado está vazio.'
        return JsonResponse(resp)

    header_row = rows[0]
    normalized_header = [_normalize_header(value) for value in header_row]
    header_map = {value: index for index,
                  value in enumerate(normalized_header) if value}

    def resolve_index(possible_keys):
        for key in possible_keys:
            if key in header_map:
                return header_map[key]
        return None

    code_idx = resolve_index(
        ['codigo', 'código', 'codigo do produto', 'código do produto', 'sku'])
    name_idx = resolve_index(['nome', 'produto', 'nome do produto'])
    description_idx = resolve_index(['descricao', 'descrição', 'detalhes'])
    category_idx = resolve_index(['categoria', 'categoria do produto'])
    price_idx = resolve_index(['preco', 'preço', 'valor', 'valor de venda'])
    cost_idx = resolve_index(['custo', 'custo unitario', 'custo unitário'])
    status_idx = resolve_index(['status'])

    if None in (code_idx, name_idx, category_idx, status_idx):
        resp['msg'] = 'Cabeçalho inválido. Utilize o modelo de importação disponibilizado.'
        return JsonResponse(resp)

    created_count = 0
    updated_count = 0
    error_rows = []

    def is_empty_row(row):
        return all(
            (
                cell is None
                or (isinstance(cell, str) and not cell.strip())
                or (not isinstance(cell, str) and str(cell).strip() == '')
            )
            for cell in row
        )

    for row_number, row in enumerate(rows[1:], start=2):
        if not row or is_empty_row(row):
            continue

        code_cell = row[code_idx] if len(row) > code_idx else None
        name_cell = row[name_idx] if len(row) > name_idx else None
        description_cell = (
            row[description_idx] if description_idx is not None and len(
                row) > description_idx else None
        )
        category_cell = row[category_idx] if len(row) > category_idx else None
        price_cell = row[price_idx] if price_idx is not None and len(
            row) > price_idx else None
        cost_cell = row[cost_idx] if cost_idx is not None and len(
            row) > cost_idx else None
        status_cell = row[status_idx] if len(row) > status_idx else None

        code = str(code_cell or '').strip()
        name = str(name_cell or '').strip()
        description = '' if description_cell in (
            None, 'None') else str(description_cell or '').strip()
        category_name = str(category_cell or '').strip()

        if not code:
            error_rows.append(
                f'Linha {row_number}: o código do produto é obrigatório.')
            continue

        if not name:
            error_rows.append(
                f'Linha {row_number}: o nome do produto é obrigatório.')
            continue

        if not category_name:
            error_rows.append(
                f'Linha {row_number}: informe a categoria do produto.')
            continue

        status_value = _parse_status_cell(status_cell)
        if status_value is None:
            error_rows.append(
                f"Linha {row_number}: status inválido. Utilize 'Ativo' ou 'Inativo' (ou 1/0)."
            )
            continue

        price_value = _parse_decimal_cell(price_cell)
        cost_value = _parse_decimal_cell(cost_cell)

        category = Category.objects.filter(
            company=user_company, name__iexact=category_name).first()
        if not category:
            category = Category.objects.create(
                company=user_company,
                name=category_name,
                description='',
                status=1,
            )

        try:
            product = Products.objects.filter(
                company=user_company, code__iexact=code).first()
            if product:
                product.code = code
                product.name = name
                product.description = description
                product.category_id = category
                if price_value is not None:
                    product.price = float(price_value)
                elif product.price is None:
                    product.price = 0.0
                if cost_value is not None:
                    product.custo = float(cost_value)
                elif product.custo is None:
                    product.custo = 0.0
                product.status = status_value
                product.save()
                updated_count += 1
            else:
                product_defaults = {
                    'company': user_company,
                    'code': code,
                    'name': name,
                    'description': description,
                    'category_id': category,
                    'price': float(price_value) if price_value is not None else 0.0,
                    'custo': float(cost_value) if cost_value is not None else 0.0,
                    'status': status_value,
                }
                Products.objects.create(**product_defaults)
                created_count += 1
        except Exception as exc:
            error_rows.append(
                f'Linha {row_number}: erro ao salvar produto ({exc}).')
            continue

    if created_count or updated_count:
        if error_rows:
            messages.warning(
                request,
                (
                    f'Importação concluída: {created_count} produto(s) adicionados e '
                    f'{updated_count} atualizados. Algumas linhas foram ignoradas.'
                ),
            )
        else:
            messages.success(
                request,
                (
                    f'Importação concluída: {created_count} produto(s) adicionados e '
                    f'{updated_count} atualizados.'
                ),
            )
    else:
        if error_rows:
            messages.error(
                request,
                'Nenhum produto foi importado. Verifique o arquivo enviado.',
            )

    if error_rows:
        resp['status'] = 'partial'
        resp['errors'] = error_rows
        if created_count or updated_count:
            resp['msg'] = (
                'Importação concluída com pendências. '
                f'{created_count} produto(s) adicionados e '
                f'{updated_count} atualizados.'
            )
        else:
            resp['msg'] = (
                'Nenhum produto foi importado. Revise as pendências indicadas no arquivo.'
            )
    else:
        resp['status'] = 'success'
        resp['msg'] = (
            'Produtos importados com sucesso. '
            f'{created_count} adicionados e {updated_count} atualizados.'
        )

    resp['created'] = created_count
    resp['updated'] = updated_count

    return JsonResponse(resp)


@login_required
def download_product_template(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Produtos'

    headers = [
        'Código',
        'Nome',
        'Descrição',
        'Categoria',
        'Preço',
        'Custo',
        'Status',
    ]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    worksheet.append([
        'PRD-0001',
        'Refrigerante 350ml',
        'Lata de refrigerante padrão',
        'Bebidas',
        6.5,
        3.2,
        'Ativo',
    ])
    worksheet.append([
        'PRD-0002',
        'Pizza Família',
        'Pizza grande com 8 fatias',
        'Alimentos',
        45.0,
        22.5,
        'Inativo',
    ])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return FileResponse(
        buffer,
        as_attachment=True,
        filename='modelo_importacao_produtos.xlsx',
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


class ProductsXMLUploadView(View):
    template_name = 'catalog/products_xml_upload.html'

    def get(self, request):
        user_company = get_user_company(request)
        if not user_company:
            messages.error(
                request, 'UsuÃ¡rio nÃ£o estÃ¡ associado a nenhuma empresa.')
            return redirect('product-page')

        return render(
            request,
            self.template_name,
            {
                'expected_fields': ['cProd', 'xProd', 'vUnCom', 'vProd', 'vUnTrib'],
                'csrf_token_value': request.META.get('CSRF_COOKIE', ''),
            },
        )

    def post(self, request):
        user_company = get_user_company(request)
        if not user_company:
            return JsonResponse(
                {'status': 'failed',
                 'msg': 'UsuÃ¡rio nÃ£o estÃ¡ associado a nenhuma empresa.'}
            )

        content_type = request.META.get('CONTENT_TYPE', '')
        if content_type.startswith('application/json'):
            try:
                payload = json.loads(request.body.decode('utf-8'))
                items = payload.get('items') or []
            except (TypeError, ValueError):
                return JsonResponse(
                    {'status': 'failed', 'msg': 'Payload JSON invÃ¡lido.'}
                )
            return self._apply_items(user_company, items)

        upload_file = request.FILES.get('file')
        if not upload_file:
            return JsonResponse(
                {'status': 'failed', 'msg': 'Envie um arquivo XML para importar.'}
            )

        parsed_items, errors = self._parse_xml_items(upload_file, user_company)
        status = 'success' if parsed_items else 'failed'
        if errors and parsed_items:
            status = 'partial'
        if not parsed_items and not errors:
            errors.append('Nenhum produto encontrado no XML informado.')

        return JsonResponse(
            {
                'status': status,
                'items': parsed_items,
                'errors': errors,
                'msg': 'PrÃ©-visualizaÃ§Ã£o pronta.' if parsed_items else 'Nenhum item encontrado.',
            }
        )

    def _parse_xml_items(self, upload_file, company):
        try:
            tree = ET.parse(upload_file)
            root = tree.getroot()
        except Exception:
            return [], ['NÃ£o foi possÃ­vel ler o XML enviado.']

        def strip_tag(tag):
            return tag.split('}', 1)[-1] if '}' in tag else tag

        def find_child_text(element, name):
            for child in element:
                if strip_tag(child.tag) == name:
                    return (child.text or '').strip()
            return ''

        items = []
        errors = []
        for idx, node in enumerate(root.iter(), start=1):
            if strip_tag(node.tag) != 'det':
                continue

            prod = next(
                (child for child in node if strip_tag(child.tag) == 'prod'),
                None,
            )
            if prod is None:
                continue

            code = find_child_text(prod, 'cProd')
            name = find_child_text(prod, 'xProd')
            if not code and not name:
                continue

            xml_unit_price = _parse_decimal_cell(
                find_child_text(prod, 'vUnCom') or find_child_text(prod, 'vProd')
            )
            if xml_unit_price is None:
                xml_unit_price = _parse_decimal_cell(
                    find_child_text(prod, 'vUnTrib')
                )

            product = Products.objects.filter(
                company=company, code__iexact=code).first() if code else None
            category_name = ''
            if product and product.category_id:
                category_name = product.category_id.name

            if xml_unit_price is None:
                errors.append(
                    f'Item {idx}: nÃ£o foi possÃ­vel identificar o preÃ§o no XML para o produto {code or name}.'
                )

            price_value = xml_unit_price if xml_unit_price is not None else _parse_decimal_cell(
                getattr(product, 'price', None)
            )
            if price_value is None:
                price_value = Decimal('0')

            items.append(
                {
                    'code': code,
                    'name': name,
                    'quantity': 0,
                    'price': float(price_value),
                    'cost': float(price_value),
                    'category': category_name,
                    'status': 1,
                }
            )

        if not items:
            errors.append('Nenhum produto encontrado no XML informado.')

        return items, errors

    def _apply_items(self, company, items):
        if not isinstance(items, list):
            return JsonResponse({'status': 'failed', 'msg': 'Lista de itens invÃ¡lida.'})

        created = 0
        updated = 0
        errors: list[str] = []

        for idx, item in enumerate(items, start=1):
            code = str(item.get('code') or '').strip()
            name = str(item.get('name') or '').strip()
            if not code:
                errors.append(f'Linha {idx}: informe o cÃ³digo do produto.')
                continue
            if not name:
                errors.append(f'Linha {idx}: informe o nome do produto.')
                continue

            status_value = _parse_status_cell(item.get('status'))
            if status_value is None:
                status_value = 1

            price_value = _parse_decimal_cell(item.get('price'))
            cost_value = _parse_decimal_cell(item.get('cost'))
            if price_value is None:
                errors.append(f'Linha {idx}: informe um preÃ§o vÃ¡lido.')
                continue
            if cost_value is None:
                errors.append(f'Linha {idx}: informe um custo vÃ¡lido.')
                continue

            product = Products.objects.filter(
                company=company, code__iexact=code).first()

            category = None
            category_name = str(item.get('category') or '').strip()
            if category_name:
                category = Category.objects.filter(
                    company=company, name__iexact=category_name
                ).first()
                if not category:
                    category = Category.objects.create(
                        company=company,
                        name=category_name,
                        description='',
                        status=1,
                    )
            elif product and product.category_id:
                category = product.category_id

            if not category:
                errors.append(
                    f'Linha {idx}: informe a categoria do produto {code}.'
                )
                continue

            try:
                if product:
                    product.name = name
                    product.category_id = category
                    product.price = float(price_value)
                    product.custo = float(cost_value)
                    product.status = status_value if status_value in (0, 1) else product.status
                    product.save()
                    updated += 1
                else:
                    Products.objects.create(
                        company=company,
                        code=code,
                        name=name,
                        description='',
                        category_id=category,
                        price=float(price_value),
                        custo=float(cost_value),
                        status=status_value if status_value in (0, 1) else 1,
                    )
                    created += 1
            except Exception as exc:
                errors.append(
                    f'Linha {idx}: erro ao salvar produto ({exc}).')
                continue

        status = 'success' if not errors else 'partial'
        msg = f'Produtos atualizados: {created} criado(s) e {updated} atualizado(s).'
        if not created and not updated:
            status = 'failed'
            msg = 'Nenhum produto foi aplicado.'

        return JsonResponse(
            {
                'status': status,
                'created': created,
                'updated': updated,
                'errors': errors,
                'msg': msg,
            }
        )


@login_required
def products(request):
    user_company = get_user_company(request)
    if not user_company:
        messages.error(
            request, 'Usuário não está associado a nenhuma empresa.')
        return redirect('home-page')

    query = request.GET.get('q', '').strip()
    category_filter = request.GET.get('category', '').strip()
    status_filter = request.GET.get('status', '').strip()
    page = request.GET.get('page', 1)

    base_qs = Products.objects.filter(
        company=user_company).select_related('category_id')
    if query:
        base_qs = base_qs.filter(
            Q(name__icontains=query) | Q(code__icontains=query))
    if category_filter.isnumeric():
        base_qs = base_qs.filter(category_id=int(category_filter))
    if status_filter in ['0', '1']:
        base_qs = base_qs.filter(status=int(status_filter))

    products_qs = base_qs.order_by('-id')
    paginator = Paginator(products_qs, 20)
    try:
        products_paginated = paginator.page(page)
    except PageNotAnInteger:
        products_paginated = paginator.page(1)
    except EmptyPage:
        products_paginated = paginator.page(paginator.num_pages)

    categories = Category.objects.filter(
        status=1, company=user_company).order_by('name')
    context = {
        'page_title': 'Lista de Produtos',
        'products': products_paginated,
        'q': query,
        'category_filter': category_filter,
        'status_filter': status_filter,
        'categories': categories,
        'total_products': base_qs.count(),
        'active_products': base_qs.filter(status=1).count(),
        'inactive_products': base_qs.filter(status=0).count(),
        'avg_price': float(base_qs.aggregate(avg=Avg('price'))['avg'] or Decimal('0')),
        'total_inventory_value': float(
            base_qs.filter(status=1).aggregate(
                total=Sum('price'))['total'] or Decimal('0')
        ),
    }
    return render(request, 'catalog/products.html', context)


@login_required
def manage_products(request):
    user_company = get_user_company(request)
    if not user_company:
        messages.error(
            request, 'Usuário não está associado a nenhuma empresa.')
        return redirect('product-page')

    product = None
    combo_items = []
    categories = Category.objects.filter(status=1, company=user_company).all()
    product_id = request.GET.get('id', '').strip()
    if product_id.isnumeric() and int(product_id) > 0:
        product = (
            Products.objects.filter(id=product_id, company=user_company)
            .prefetch_related('combo_items__component')
            .first()
        )
        if product:
            combo_items = [
                {
                    'component_id': item.component_id,
                    'component_name': item.component.name,
                    'quantity': float(item.quantity),
                }
                for item in product.combo_items.all()
            ]

    component_products = (
        Products.objects.filter(company=user_company, status=1, is_combo=False)
        .order_by('name')
        .all()
    )

    context = {
        'product': product,
        'categories': categories,
        'component_products': component_products,
        'combo_items': combo_items,
    }
    return render(request, 'catalog/manage_product.html', context)


def test(request):
    return render(request, 'catalog/test.html', {'categories': Category.objects.all()})


@login_required
def save_product(request):
    data = request.POST
    resp = {'status': 'failed'}
    user_company = get_user_company(request)
    if not user_company:
        resp['msg'] = 'Usuário não está associado a nenhuma empresa.'
        return HttpResponse(json.dumps(resp), content_type='application/json')

    product_id = data.get('id', '')
    code = data.get('code', '')
    if product_id.isnumeric() and int(product_id) > 0:
        duplicate_qs = Products.objects.exclude(
            id=product_id).filter(code=code, company=user_company)
    else:
        duplicate_qs = Products.objects.filter(code=code, company=user_company)

    if duplicate_qs.exists():
        resp['msg'] = 'Código de produto já existe no banco de dados'
        return HttpResponse(json.dumps(resp), content_type='application/json')

    price = data.get('price', '0').replace(',', '.')
    cost = data.get('custo', '0').replace(',', '.')
    category = Category.objects.filter(id=data.get(
        'category_id'), company=user_company).first()
    if not category:
        resp['msg'] = 'Categoria não encontrada ou não pertence à sua empresa'
        return HttpResponse(json.dumps(resp), content_type='application/json')

    is_combo = data.get('is_combo') in ('1', 'true', 'on')
    combo_total_quantity = (data.get('combo_total_quantity') or '').strip()
    combo_max_flavors = (data.get('combo_max_flavors') or '').strip()

    try:
        price_value = float(price)
    except ValueError:
        resp['msg'] = 'Informe um preço válido.'
        return HttpResponse(json.dumps(resp), content_type='application/json')

    try:
        cost_value = float(cost)
    except ValueError:
        resp['msg'] = 'Informe um custo válido.'
        return HttpResponse(json.dumps(resp), content_type='application/json')

    product_instance = None
    if product_id.isnumeric() and int(product_id) > 0:
        product_instance = Products.objects.filter(
            id=product_id, company=user_company).first()
        if not product_instance:
            resp['msg'] = 'Produto não encontrado.'
            return HttpResponse(json.dumps(resp), content_type='application/json')
    else:
        product_instance = Products(company=user_company)

    product_instance.code = code
    product_instance.category_id = category
    product_instance.name = data.get('name', '')
    product_instance.description = data.get('description', '')
    product_instance.price = price_value
    product_instance.custo = cost_value
    product_instance.status = int(data.get('status', 1))
    product_instance.is_combo = is_combo

    if is_combo and combo_total_quantity:
        try:
            product_instance.combo_total_quantity = Decimal(
                combo_total_quantity.replace(',', '.'))
        except Exception:
            resp['msg'] = 'Informe uma quantidade total válida para o combo.'
            return HttpResponse(json.dumps(resp), content_type='application/json')
    else:
        product_instance.combo_total_quantity = None

    if is_combo and combo_max_flavors:
        if combo_max_flavors.isnumeric():
            product_instance.combo_max_flavors = int(combo_max_flavors)
        else:
            resp['msg'] = 'Informe um número válido de sabores máximos.'
            return HttpResponse(json.dumps(resp), content_type='application/json')
    else:
        product_instance.combo_max_flavors = None

    try:
        product_instance.save()
    except Exception as exc:
        resp['msg'] = str(exc)
        return HttpResponse(json.dumps(resp), content_type='application/json')

    if is_combo:
        component_ids = data.getlist('combo_component_id[]')
        component_qtys = data.getlist('combo_component_qty[]')
        components_payload = {}

        for comp_id, qty_str in zip(component_ids, component_qtys):
            comp_id = (comp_id or '').strip()
            if not comp_id:
                continue

            try:
                component_obj = Products.objects.get(
                    id=int(comp_id), company=user_company
                )
            except (Products.DoesNotExist, ValueError):
                resp['msg'] = 'Selecione componentes válidos para o combo.'
                return HttpResponse(json.dumps(resp), content_type='application/json')

            if component_obj.id == product_instance.id:
                resp['msg'] = 'O combo não pode incluir ele mesmo como componente.'
                return HttpResponse(json.dumps(resp), content_type='application/json')

            if component_obj.is_combo:
                resp['msg'] = 'Não é possível utilizar outro combo como componente.'
                return HttpResponse(json.dumps(resp), content_type='application/json')

            qty_clean = (qty_str or '0').replace(',', '.')
            try:
                qty_value = Decimal(qty_clean)
            except Exception:
                resp['msg'] = 'Informe quantidades válidas para os componentes.'
                return HttpResponse(json.dumps(resp), content_type='application/json')

            if qty_value < 0:
                resp['msg'] = 'As quantidades dos componentes não podem ser negativas.'
                return HttpResponse(json.dumps(resp), content_type='application/json')

            payload = components_payload.setdefault(
                component_obj.id,
                {'component': component_obj, 'quantity': Decimal('0')},
            )
            payload['quantity'] += qty_value

        if not components_payload:
            resp['msg'] = 'Cadastre pelo menos um componente para o combo.'
            return HttpResponse(json.dumps(resp), content_type='application/json')

        ProductComboItem.objects.filter(combo=product_instance).delete()
        for payload in components_payload.values():
            ProductComboItem.objects.create(
                combo=product_instance,
                component=payload['component'],
                quantity=payload['quantity'],
                company=user_company,
            )
    else:
        ProductComboItem.objects.filter(combo=product_instance).delete()

    resp['status'] = 'success'
    messages.success(request, 'Produto salvo com sucesso.')
    return HttpResponse(json.dumps(resp), content_type='application/json')


@login_required
def delete_product(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'failed', 'msg': 'Método inválido.'})

    data = request.POST
    resp = {'status': 'failed'}
    user_company = get_user_company(request)
    if not user_company:
        resp['msg'] = 'Usuário não está associado a nenhuma empresa.'
        return JsonResponse(resp)

    product_id = data.get('id', '').strip()
    if not product_id.isdigit():
        resp['msg'] = 'Identificador de produto inválido.'
        return JsonResponse(resp)

    product = Products.objects.filter(
        id=int(product_id), company=user_company).first()
    if not product:
        resp['msg'] = 'Produto não encontrado para sua empresa.'
        return JsonResponse(resp)

    try:
        product.delete()
    except ProtectedError:
        resp['msg'] = (
            'Não é possível deletar o produto porque ele está vinculado a outros registros.'
        )
        return JsonResponse(resp)

    messages.success(request, 'Produto deletado com sucesso.')
    return JsonResponse({'status': 'success'})
